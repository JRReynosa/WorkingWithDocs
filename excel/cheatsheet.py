import openpyxl

# Import workbook
wb = openpyxl.load_workbook('example.xlsx')

# The workbook's sheets' names.
wb.sheetnames
# Output: ['Sheet1', 'Sheet2', 'Sheet3']

# Get a sheet from the workbook.
sheet = wb['Sheet3']
sheet
# Output: <Worksheet "Sheet3">
type(sheet)
# Output: <class 'openpyxl.worksheet.worksheet.Worksheet'>

# Get the sheet's title as a string.
sheet.title
# Output: 'Sheet3'

# Get the active sheet.
anotherSheet = wb.active
anotherSheet
# Output: <Worksheet "Sheet1">

# Get a cell from the sheet.
sheet['A1']
# Output: <Cell 'Sheet1'.A1>

# Get the value from the cell.
sheet['A1'].value
# Output: datetime.datetime(2015, 4, 5, 13, 34, 2)

# Get another cell from the sheet.
c = sheet['B1']
c.value
# Output: 'Apples'

# Edit the cell's value.
sheet['A1'] = 'Hello, world!'

# Get the row, column, and value from the cell.
'Row %s, Column %s is %s' % (c.row, c.column, c.value)
# Output: 'Row 1, Column B is Apples'
'Cell %s is %s' % (c.coordinate, c.value)
# Output: 'Cell B1 is Apples'

# Get cell by row & col coordinates
sheet.cell(row=1, column=2)
# Output: <Cell 'Sheet1'.B1>

sheet.cell(row=1, column=2).value
# Output: 'Apples'

# Go through every other row:
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
# Output:
# 1 Apples
# 3 Pears
# 5 Apples
# 7 Strawberries

# Get second column's cells.
list(sheet.columns)[1]
# Output:
# (<Cell 'Sheet1'.B1>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.
# B4>, <Cell 'Sheet1'.B5>, <Cell 'Sheet1'.B6>, <Cell 'Sheet1'.B7>)

for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
# Output:
# Apples
# Cherries
# Pears
# Oranges
# Apples
# Bananas
# Strawberries

# -------------------------------------------------------------------
# Working with Workbooks
# Create a blank workbook.
# It starts with one sheet.
wb = openpyxl.Workbook()
wb.sheetnames
# Output: ['Sheet']

sheet = wb.active
sheet.title
# Output: 'Sheet'

# Change title.
sheet.title = 'Spam Bacon Eggs Sheet'
wb.sheetnames
# Output: ['Spam Bacon Eggs Sheet']

# Save the workbook.
wb.save('example_copy.xlsx')

# Add a new sheet.
wb.create_sheet()
# Output: <Worksheet "Sheet1">

wb.sheetnames
# Output: ['Sheet', 'Sheet1']

# Create a new sheet at index 0.
wb.create_sheet(index=0, title='First Sheet')
# Output: <Worksheet "First Sheet">

wb.sheetnames
# Output: ['First Sheet', 'Sheet', 'Sheet1']

wb.create_sheet(index=2, title='Middle Sheet')
# Output: <Worksheet "Middle Sheet">

wb.sheetnames
# Output: ['First Sheet', 'Sheet', 'Middle Sheet', 'Sheet1']

del wb['Middle Sheet']
del wb['Sheet1']
wb.sheetnames
# Output: ['First Sheet', 'Sheet']
